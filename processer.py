import os
from pathlib import Path

from openpyxl import load_workbook

from models import WGrid, Grid500, Grid100, Grid50, Grid20

class PTSFileProcessor:
    GRIDS = {'1000': WGrid,
        '500': Grid500,
        '100': Grid100,
        '50': Grid50,
        '20': Grid20}

    def __init__(self, import_orgainser, controller):
        self.import_orgainser = import_orgainser
        self.controller = controller
     
    def process_pts_files(self, directory, grid_size, description, survey_type, survey_date, end_date):
        self.survey_date = survey_date.strftime('%y%m%d')
        self.end_date = end_date
        if self.end_date is not None:
            self.end_date = self.end_date.strftime('%y%m%d')
        
        count = {}
        GRID_CLASS = self.GRIDS[grid_size]
        for imported_file in self.import_orgainser:
            if imported_file.exists and imported_file.file_active:
                with open(imported_file.path, 'r', encoding='utf-8-sig') as infile:
                    for line in infile:
                        s = line.split()
                        if len(s) > 2:
                            e, n = float(s[0]), float(s[1])
                            wgrid = self.controller.get_wgrid(e,n)
                            if grid_size != '1000':
                                grid_cords = wgrid.calculate_sub_grid(e, n, grid_size)
                            else:
                                grid_cords = (wgrid.lower_x, wgrid.lower_y)
                            if grid_cords not in count:
                                count[grid_cords] = 1
                            else:
                                count[grid_cords] += 1 
        try:
            outfiles = {}
            file_num = 1
            for imported_file in self.import_orgainser:
                if imported_file.exists and imported_file.file_active:
                    with open(imported_file.path, 'r', encoding='utf-8-sig') as infile:               
                        for line in infile:
                            s = line.split()
                            if len(s) > 2:
                                e, n = float(s[0]), float(s[1])
                                wgrid = self.controller.get_wgrid(e,n)
                                if grid_size != '1000':
                                    grid_cords = wgrid.calculate_sub_grid(e, n, grid_size)
                                    file_name = f"HS2-{wgrid.name}-{GRID_CLASS.sub_grid_name(str(grid_cords[0]),str(grid_cords[1]))}-{self.get_survey_date()}{'-'+ description}"
                                else:
                                    grid_cords = (wgrid.lower_x, wgrid.lower_y)
                                    file_name = f"HS2-{wgrid.name}-{self.get_survey_date()}{'-'+ description}"
                                    
                                if grid_cords not in outfiles:
                                    file_name += f"({file_num}of{len(count)}){'-' + survey_type}.pts"
                                    file_name = os.path.join(directory, file_name)
                                    outfiles[grid_cords] = open(file_name, 'w')
                                    outfiles[grid_cords].write(str(count[grid_cords])+ '\n')
                                    file_num += 1
                                outfiles[grid_cords].write(self.get_line(s))
                    
        finally:
            for outfile in outfiles.values():
                outfile.close()

    def get_survey_date(self):
        if self.end_date is None:
            return self.survey_date
        else:
            return f'{self.survey_date}-{self.end_date}'
        
    def get_line(self, s):
        line = ''
        for l in s: 
            line += l + ' '
        line += '\n'
        return line 
                         
    def check_pts_line(self, pts_line):
        if len(pts_line) > 2:
            return True 
        return False   
         
                 
class GridCreator:
    """Class to create internal w and sub grid files"""
    GRIDS = {'1000': WGrid,
        '500': Grid500,
        '100': Grid100,
        '50': Grid50,
        '20': Grid20}
    
    FILE_TYPE = '.pts'

    def __init__(self, grid_size):
        self.GRID_CLASS = self.GRIDS.get(grid_size)
        self.grid_files = {}  # {} WGrid: {(EN Lower left corder ):File name}}

    def sub_grid_file(self, wgrid, e, n):
        """Passed in wgrid and lower e, n cords 
        returns the opened subgrid"""
        grid = self.grid_files.setdefault(wgrid, {})
        return grid.setdefault((e, n), 
            open(f'HS2-{wgrid.name}-{self.GRID_CLASS.sub_grid_name(str(e),str(n))}{self.FILE_TYPE}', 'w'))

    def w_grid_file(self, wgrid):
        """Passed in wgrid returns the opened wgrid file"""
        grid = self.grid_files.setdefault(wgrid, {})
        return grid.setdefault((wgrid.lower_x, wgrid.lower_y), open(f'HS2-{wgrid.name}{self.FILE_TYPE}', 'w'))

    def close_grid_files(self):
         """closes up the grid files on exit"""
         ((f.close() for f in v.values()) for v in self.grid_files.values())


class ProjectWiseWritter:
    def __init__(self, pod_filenames, spreadsheet_filename, drawn_by,
        issue_purpose, drawn_date):
        self.spreadsheet_filename = spreadsheet_filename
        self.pod_filenames = pod_filenames
        self.workbook = load_workbook(self.spreadsheet_filename)
        self.sheet = self.workbook.active
        self.drawn_by = drawn_by
        self.issue_purpose = issue_purpose
        self.drawn_date = drawn_date.strftime('%d/%m/%y')

    def check_pod_names(self):
        """check pod names are in correct format"""
        for pod in self.pod_filenames:
            if pod.exists and pod.file_active: 
                result = self._pod_num(pod.file_name)
                if result is None:
                    return False 
        return True
       
    def check_excel(self):
        """check correct number of rows in excel"""
        return len(self.pod_filenames) == (self.sheet.max_row -1)
    
    def process(self):
        for n, row in enumerate(self.sheet.iter_rows()):
            n = n+1
            if n != 1: 
                n = str(n)
                doc_str = self.sheet['E'+ n].value
                pod_file = self.get_pod_file(
                    self.get_doc_number(doc_str))

                if pod_file is not None:
                    self.sheet['A'+ n] = pod_file.path.stem
                    self.sheet['B' + n] = self.sheet['P' + n].value + '.pod'
                    self.sheet['C' + n] = pod_file.normpath
                    self.sheet['E' + n] = self.sheet['P' + n].value 
                    self.sheet['Q' + n] = 'Survey Data File'
                    self.sheet['R' + n] = self.drawn_by.title()
                    self.sheet['U' + n] = self.issue_purpose
                    self.sheet['AG' + n] = self.sheet['J' + n].value
                    self.sheet['AJ' + n] = 'NOT APPLICABLE'
                    self.sheet['AP' + n] = self.drawn_date
                    self.sheet['AR' + n] = 'Construction'
                    title_1 = self.sheet['AG' + n].valuen
                    title_2 = 'Point Cloud File'
                    title_3 = self.title_3(pod_file)
                    self.sheet['AS' + n] = f'{title_1}, {title_2}, {title_3}'
                    self.sheet['AT' + n] = title_1
                    self.sheet['AU' + n] = title_2
                    self.sheet['AV' + n] = title_3        
        
        self.workbook.save(self.spreadsheet_filename)
       
    def title_3(self, filename):
        s = filename.path.stem.split('-')
        return f'{s[-2]}-{s[-1]}'

    def _pod_num(self, filename):
        try:
            spl = filename.split(')-')
            s = spl[0].split('(')
            return int(s[-1].split('of')[0])
        except:
            return 

    def get_pod_file(self, doc_num):
        for pod in self.pod_filenames:
            if pod.exists and pod.file_active:
                pod_num = self._pod_num(pod.file_name)
                if doc_num == pod_num:
                    return pod

    def get_doc_number(self, doc_str):
        for n, l in enumerate(doc_str):
            try:
                num = int(l)
                if num > 0:
                    return int(doc_str[n:])
            except ValueError: 
                continue 